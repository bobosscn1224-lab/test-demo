"""Report Builder — copy template, fill activities, LLM-generate summary."""

from __future__ import annotations

import copy
import logging
import os
import re
from datetime import datetime

logger = logging.getLogger(__name__)


async def build_report_xlsx(
    template_path: str,
    output_path: str,
    start_date: str,
    end_date: str,
    activities: dict[int, list[dict]],
) -> str:
    """Build a weekly report Excel using the template format.

    Flow (same as weekly_report skill):
      1. Copy the latest template sheet → preserves all formatting
      2. Clear old data from the new sheet
      3. Write activity data to D/E columns
      4. LLM-generate summary text for B/C columns
      5. Fill empty cells with "—"
    """
    import openpyxl
    from app.services.llm_service import llm_service

    s = datetime.strptime(start_date, "%Y-%m-%d")
    e = datetime.strptime(end_date, "%Y-%m-%d")
    sheet_name = f"{s.month}.{s.day}-{e.month}.{e.day}"

    # ── 1. Load template and copy latest sheet ──
    wb = openpyxl.load_workbook(template_path)

    # Find template source sheet
    src_sheet = None
    date_sheets = []
    for name in wb.sheetnames:
        if re.match(r'^\d+\.\d+-\d+\.\d+$', name):
            m = re.match(r'^(\d+)\.(\d+)-', name)
            if m:
                month, day = int(m.group(1)), int(m.group(2))
                date_sheets.append(((month, day), name))
    if date_sheets:
        date_sheets.sort()
        src_sheet = date_sheets[-1][1]
    elif wb.sheetnames:
        # Fallback: first non-default sheet
        for name in wb.sheetnames:
            if name not in ("Sheet1", "sheet1"):
                src_sheet = name
                break
        if not src_sheet:
            src_sheet = wb.sheetnames[0]

    if not src_sheet:
        wb.close()
        raise ValueError("模板文件中没有可用的sheet")

    # Remove existing sheet with same name
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Copy template sheet
    _copy_sheet(wb, src_sheet, sheet_name)

    # ── 2. Remove old data ──
    ws = wb[sheet_name]
    # Clear old D/E columns (keep A/B/C structure)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=5):
        for cell in row:
            cell.value = None
    # Clear placeholder text in summary rows
    placeholder_patterns = ["由AI根据", "本周计划与目标", "主要进展和成果", "下周计划与目标"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for pat in placeholder_patterns:
                    if pat in cell.value:
                        cell.value = None
                        break

    # ── 3. Write activity data to D/E columns ──
    day_names = ["周一", "周二", "周三", "周四", "周五"]
    row = 2
    day_boundaries = {}  # day_idx -> (start_row, end_row)

    for day_idx in range(5):
        day_activities = activities.get(day_idx, [])
        d = s + __import__('datetime').timedelta(days=day_idx)
        date_str = d.strftime("%Y-%m-%d")
        day_label = day_names[day_idx]

        # Sort morning first, then by time
        sorted_acts = sorted(day_activities, key=lambda a: (
            0 if a.get('period') == '上午' else 1,
            a.get('time_start', '') or ''
        ))

        first_row = row
        for act in sorted_acts:
            period = act.get('period', '')
            time_start = act.get('time_start', '') or ''
            time_end = act.get('time_end', '') or ''

            # Default times within work hours
            if period == '上午':
                if not time_start or time_start < '09:00':
                    time_start = '09:00'
            elif period == '下午':
                if not time_start or time_start < '14:00':
                    time_start = '14:00'

            time_label = f"{time_start}-{time_end}" if time_end else time_start
            activity = act.get('activity', '') or ''
            result = act.get('result', '') or '—'

            ws.cell(row, 1, value=f"{d.month}月{d.day}日 {day_label}")
            ws.cell(row, 2, value=period)
            ws.cell(row, 3, value=time_label)
            ws.cell(row, 4, value=activity)
            ws.cell(row, 5, value=result)
            row += 1

        if row > first_row + 1:
            day_boundaries[day_idx] = (first_row, row - 1)
            # Merge date column for this day
            try:
                ws.merge_cells(None, first_row, 1, row - 1, 1)
            except Exception:
                pass
        elif row > first_row:
            day_boundaries[day_idx] = (first_row, row - 1)

    # ── 4. Build activity summary text for LLM ──
    activity_lines = []
    for day_idx in range(5):
        acts = activities.get(day_idx, [])
        if acts:
            d_label = day_names[day_idx]
            for a in acts:
                activity_lines.append(f"{d_label} {a.get('period','')}: {a.get('activity','')} → {a.get('result','')}")

    # ── 5. LLM-generate summary ──
    summary_b, summary_c = "", ""
    if activity_lines:
        try:
            from app.skills.weekly_report.llm_ops import call_llm_for_json
            summary_prompt = f"""请根据本周工作内容，生成两个简短的周报总结：

本周工作内容：
{chr(10).join(activity_lines[:30])}

请输出JSON格式：
[{{"cell":"Bx","value":"主要进展和成果的简练总结（2-4条，每条一行）"}},{{"cell":"Cx","value":"下周计划和目标的简练描述（2-4条，每条一行）"}}]

总结要专业简练，每条15-30字，直接写内容，不要加"本周"等前缀。"""
            summary_result = await call_llm_for_json(summary_prompt)
            for item in summary_result:
                col = str(item.get("cell", ""))[0] if item.get("cell") else ""
                val = item.get("value", "")
                if col == "B" and val:
                    summary_b += val + "\n" if summary_b else val
                elif col == "C" and val:
                    summary_c += val + "\n" if summary_c else val
        except Exception as exc:
            logger.warning("Summary generation failed: %s", exc)

    # Write summary to the B/C columns below the data
    summary_row = row + 1
    if summary_b:
        ws.cell(summary_row, 1, value="本周计划与目标")
        ws.cell(summary_row, 2, value="主要进展和成果").font = openpyxl.styles.Font(bold=True, size=11)
        ws.cell(summary_row, 3, value="下周计划与目标").font = openpyxl.styles.Font(bold=True, size=11)
        summary_row += 1
        for line in summary_b.strip().split('\n'):
            ws.cell(summary_row, 2, value=line.strip())
            summary_row += 1
        summary_row = row + 1  # reset for C column
        for line in summary_c.strip().split('\n'):
            ws.cell(summary_row, 3, value=line.strip())
            summary_row += 1

    # ── 6. Save ──
    wb.save(output_path)
    wb.close()
    logger.info("ReportBuilder: created %s with %d activity rows + summary", output_path, row - 2)
    return output_path


def _copy_sheet(wb, source_name: str, target_name: str):
    """Copy a worksheet within the same workbook, preserving formatting."""
    from openpyxl.utils import get_column_letter

    ws_source = wb[source_name]
    ws_target = wb.create_sheet(title=target_name)

    for row in ws_source.iter_rows(min_row=1, max_row=ws_source.max_row, max_col=ws_source.max_column):
        for cell in row:
            new_cell = ws_target.cell(row=cell.row, column=cell.column)
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy.copy(cell.alignment)

    for merged_range in ws_source.merged_cells.ranges:
        ws_target.merge_cells(str(merged_range))

    for col_idx in range(1, ws_source.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_source.column_dimensions:
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width

    for row_idx in range(1, ws_source.max_row + 1):
        if row_idx in ws_source.row_dimensions:
            ws_target.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height
