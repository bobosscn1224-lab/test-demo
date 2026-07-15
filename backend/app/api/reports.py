"""Weekly Report Feature API — structured activity-based generation."""
from __future__ import annotations

import json as _json
import logging
import os
import re as _re
import shutil
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
import openpyxl

from app.skills.weekly_report import excel_ops, date_utils, llm_ops
from app.skills.weekly_report.constants import DAY_NAMES, OUTPUT_DIR
from app.services._paths import PUBLIC_DIR, WEEKLY_REPORT_DIR
from app.services.report_store import add_record, list_records, delete_record as store_delete, get_record
from app.services.activity_store import load as draft_load, save as draft_save
from app.services.llm_service import llm_service
from app.services.llm_logger import log_call

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/reports", tags=["reports"])

_DOWNLOAD_DIR = str(PUBLIC_DIR)
_DATA_OUTPUT_DIR = str(WEEKLY_REPORT_DIR)

# ═══════════════════ Models ═══════════════════

class WeekOption(BaseModel):
    label: str; is_current: bool = False; offset: int; start: str; end: str; monday: str; sunday: str

class ReportListItem(BaseModel):
    filename: str; sheet_name: str; date_range: str; created_at: str = ""; download_url: str = ""

class ReportPreview(BaseModel):
    filename: str; sheet_name: str; date_range: str
    days: list[dict]; summary_b: str; summary_c: str

class ActivityJSON(BaseModel):
    day_index: int; period: str = ""; time_start: str = ""; time_end: str = ""; activity: str = ""; result: str = ""

class GenerateReportRequest(BaseModel):
    start_date: str; end_date: str
    activities: list[ActivityJSON] = Field(default_factory=list)
    template_filename: str | None = None

class GenerateReportResponse(BaseModel):
    success: bool; message: str = ""; filename: str = ""; download_url: str = ""; path: str = ""; cells_updated: int = 0; sheet_name: str = ""

class AutoFillActivity(BaseModel):
    period: str = ""; time_start: str = ""; time_end: str = ""; activity: str = ""; result: str = ""

class AutoFillDay(BaseModel):
    day_index: int; day_name: str; activities: list[AutoFillActivity] = Field(default_factory=list)

class AutoFillResponse(BaseModel):
    days: list[AutoFillDay]; warnings: list[str] = Field(default_factory=list); detected_dates: list[str] = Field(default_factory=list)

# ═══════════════════ List / Templates / Delete ═══════════════════

@router.get("/list", response_model=list[ReportListItem])
async def list_reports():
    return [ReportListItem(filename=r["filename"], sheet_name=r.get("sheet_name",""), date_range=r.get("date_range",""), created_at=r.get("created_at",""), download_url=r.get("download_url","")) for r in list_records()]

@router.get("/templates", response_model=list[ReportListItem])
async def list_templates():
    records = list_records()
    if records:
        return [ReportListItem(filename=r["filename"], sheet_name=r.get("sheet_name",""), date_range=r.get("date_range",""), created_at=r.get("created_at",""), download_url=r.get("download_url","")) for r in records]
    import glob as _glob
    results, seen = [], set()
    for d in [OUTPUT_DIR, _DOWNLOAD_DIR, _DATA_OUTPUT_DIR]:
        if not os.path.isdir(d): continue
        for fp in sorted(_glob.glob(os.path.join(d, "*.xlsx")), reverse=True):
            fn = os.path.basename(fp)
            if fn.startswith("~$") or fn in seen: continue
            seen.add(fn)
            sh = fn.replace("26年周工作总结和下周计划-ZB-", "").replace(".xlsx", "")
            results.append(ReportListItem(filename=fn, sheet_name=sh, date_range=sh, created_at="", download_url=f"/api/skills/download/{fn}"))
    return results

@router.delete("/{filename:path}")
async def delete_report(filename: str):
    import urllib.parse; filename = urllib.parse.unquote(filename)
    found, deleted = store_delete(filename)
    if not found: raise HTTPException(status_code=404, detail="记录未找到")
    return {"success": True, "deleted": len(deleted)}

# ═══════════════ Weeks ═══════════════

@router.get("/weeks", response_model=list[WeekOption])
async def list_available_weeks():
    return date_utils.build_weeks_data()

# ═══════════════ Auto-fill ═══════════════

class AutoFillRequest(BaseModel):
    text: str; start_date: str; end_date: str

@router.post("/auto-fill", response_model=AutoFillResponse)
async def auto_fill(req: AutoFillRequest):
    s = datetime.strptime(req.start_date, "%Y-%m-%d")
    prompt = """从用户输入中提取并整理本周工作内容。每项输出一条。

	用户输入：
	{}

	输出JSON数组，每项格式：
	{{"date":"X月X日","period":"上午","time_start":"09:00","time_end":"10:00","activity":"专业简洁的工作描述","result":"完成情况与进展"}}

	要求：
	- activity：将用户原文用专业简练的语言重述，保留关键信息，去掉口语化表述
	- result：根据activity合理推断完成情况并做专业表述（如「完成方案框架并与团队同步」「输出会议纪要并推动决议」）。用户提到了结果就用用户的，没提到就写1句简练的完成表述
	- time_start/time_end：用户写了就提取，没写按默认（上午09:00-10:00，下午14:00-15:00）
	- 用户没提到的日期不输出
	- 目标周({}~{})外的日期不输出""".format(req.text[:3000], req.start_date, req.end_date)
    draft_save("default", req.start_date, req.end_date, [[{"period":"","time_start":"","time_end":"","activity":req.text,"result":"用户原文中的完成情况或结果"}]])
    try:
        resp = await llm_service.chat(system_prompt="你是严格的JSON格式化助手。只输出JSON数组。", messages=[{"role":"user","content":prompt}], max_tokens=2000, temperature=0, thinking={"type":"disabled"})
        text = "";
        if resp.content:
            for b in resp.content:
                if hasattr(b,"text") and b.text: text += b.text
        text = text.strip()
        if text.startswith('```'): text = text.split('\n',1)[-1].rsplit('```',1)[0]
        log_call("auto-fill", "JSON extractor", prompt, text)
    except Exception as exc:
        log_call("auto-fill", "JSON extractor", prompt, "", str(exc))
        raise HTTPException(status_code=500, detail=f"LLM调用失败：{exc}")

    days = [AutoFillDay(day_index=i, day_name=DAY_NAMES[i]) for i in range(5)]
    detected_dates: list[str] = []
    try:
        items = _json.loads(text)
        if isinstance(items, list):
            for item in items:
                if not isinstance(item, dict): continue
                ds = str(item.get('date','')).strip()
                if ds: detected_dates.append(ds)
                period = str(item.get('period','')).strip()
                ts = str(item.get('time_start', item.get('time',''))).strip()
                te = str(item.get('time_end','')).strip()
                act = str(item.get('activity','')).strip()
                res = str(item.get('result','')).strip()
                if not act: continue
                if ts and len(ts)==4 and ':' not in ts: ts = ts[:2]+':'+ts[2:]
                if te and len(te)==4 and ':' not in te: te = te[:2]+':'+te[2:]
                no_user_time = not (item.get('time_start') or item.get('time'))
                if no_user_time:
                    if period == '上午':
                        if not ts or ts < '09:00': ts = '09:00'
                        if not te or te > '12:00': te = '10:00'
                    elif period == '下午':
                        if not ts or ts < '14:00': ts = '14:00'
                        if not te or te > '18:00': te = '15:00'
                di = -1
                for i in range(5):
                    d = s + timedelta(days=i)
                    if ds in (f"{d.month}月{d.day}日", f"{d.month}.{d.day}"): di = i; break
                if di < 0: continue
                days[di].activities.append(AutoFillActivity(period=period, time_start=ts, time_end=te, activity=act, result=res))
    except (_json.JSONDecodeError, Exception) as exc:
        logger.warning("Auto-fill parse failed: %s", exc)

    for day in days:
        if day.activities:
            day.activities.sort(key=lambda a: (a.period!='上午', a.time_start or ''))

    warnings: list[str] = []
    valid_pats = set()
    for i in range(5):
        d = s + timedelta(days=i)
        valid_pats.add(f"{d.month}月{d.day}日"); valid_pats.add(f"{d.month}.{d.day}")
    if detected_dates:
        outside = [dd for dd in detected_dates if dd not in valid_pats]
        inside = [dd for dd in detected_dates if dd in valid_pats]
        if outside and not inside:
            for d in days: d.activities = []
            warnings.append(f"日期完全不匹配：{', '.join(outside[:5])} 不在 {req.start_date}~{req.end_date}")
        elif outside:
            warnings.append(f"以下日期不在目标周内已忽略：{', '.join(outside[:5])}")
    else:
        if not any(d.activities for d in days):
            for d in days: d.activities = []
            warnings.append("未识别到日期。请用「X月X日：上午…」格式标注。")

    # Persist to JSON draft
    draft_save("default", req.start_date, req.end_date, [[{"period": a.period, "time_start": a.time_start, "time_end": a.time_end, "activity": a.activity, "result": a.result} for a in day.activities] for day in days])
    return AutoFillResponse(days=days, warnings=warnings, detected_dates=detected_dates)

# ═══════════════ Draft ═══════════════

class DraftData(BaseModel):
    session_id: str = "default"
    start_date: str = ""
    end_date: str = ""
    days: list[list[dict]] = Field(default_factory=lambda: [[] for _ in range(5)])

@router.post("/draft/save")
async def save_draft(data: DraftData):
    draft_save(data.session_id, data.start_date, data.end_date, data.days)
    return {"ok": True}

@router.get("/draft/load")
async def load_draft(session_id: str = "default"):
    d = draft_load(session_id)
    return {"start_date": d["start_date"], "end_date": d["end_date"], "days": d["days"]}

# ═══════════════ Enrich ═══════════════

class EnrichRequest(BaseModel):
    activities: list[ActivityJSON]
    start_date: str; end_date: str

@router.post("/enrich", response_model=AutoFillResponse)
async def enrich_activities(req: EnrichRequest):
    s = datetime.strptime(req.start_date, "%Y-%m-%d")
    existing = "\n".join(
        f"{['周一','周二','周三','周四','周五'][a.day_index]} {a.period} {a.time_start}-{a.time_end}: {a.activity}（{a.result}）"
        for a in req.activities if a.activity
    )
    draft_save("default", req.start_date, req.end_date,
               [[{"period": a.period, "time_start": a.time_start, "time_end": a.time_end, "activity": a.activity, "result": a.result}
                 for a in req.activities if a.day_index == i] for i in range(5)])

    blanks = [a for a in req.activities if not a.activity]
    if not blanks:
        # Enrich: fill empty result fields
        acts_without_result = [a for a in req.activities if a.activity and not a.result]
        if acts_without_result:
            prompt = "为以下工作生成完成情况（每条1句，专业简练）：\\n" + "\\n".join(f"- {a.activity}" for a in acts_without_result[:10]) + "\\n\\n输出JSON：[{\"idx\":0,\"result\":\"完成情况\"}]"
            try:
                resp = await llm_service.chat(system_prompt="你是周报完善助手。", messages=[{"role":"user","content":prompt}], max_tokens=2000, temperature=0.2, thinking={"type":"disabled"})
                t = ""
                if resp.content:
                    for b in resp.content:
                        if hasattr(b,"text") and b.text: t += b.text
                t = t.strip().lstrip().strip()
                items = _json.loads(t)
                if isinstance(items, list):
                    for item in items:
                        idx = item.get('idx', -1)
                        if 0 <= idx < len(acts_without_result):
                            acts_without_result[idx].result = str(item.get('result', ''))
            except Exception: pass
        return AutoFillResponse(days=[AutoFillDay(day_index=i, day_name=DAY_NAMES[i], activities=[
            AutoFillActivity(period=a.period, time_start=a.time_start, time_end=a.time_end, activity=a.activity, result=a.result)
            for a in req.activities if a.day_index == i
        ]) for i in range(5)], warnings=["已完善完成情况"] if acts_without_result else [], detected_dates=[])

    blank_count = len(blanks)
    blank_info = "\\n".join(f"第{i+1}行 {['周一','周二','周三','周四','周五'][a.day_index]} {a.period or '上午'} {a.time_start or '?'}-{a.time_end or '?'}: 请填入" for i, a in enumerate(blanks))
    existing_info = "\\n".join(f"{['周一','周二','周三','周四','周五'][a.day_index]} {a.period} {a.time_start}-{a.time_end}: {a.activity}" for a in req.activities if a.activity)

    prompt = f"""根据已有日程上下文，为{blank_count}个空白时段填写合理的工作内容。

已有日程：
{existing_info[:1500]}

需填充（{blank_count}个）：
{blank_info}

要求：每条15-30字，专业简练，结合上下文推断。目标周({req.start_date}~{req.end_date})。
输出JSON：[{{"date":"X月X日","period":"上午","time_start":"09:00","time_end":"10:00","activity":"描述","result":"完成情况"}}]"""

    try:
        resp = await llm_service.chat(system_prompt="你是周报完善助手。输出JSON数组。", messages=[{"role":"user","content":prompt}], max_tokens=3000, temperature=0.2, thinking={"type":"disabled"})
        t = ""
        if resp.content:
            for b in resp.content:
                if hasattr(b,"text") and b.text: t += b.text
        t = t.strip()
        if t.startswith('```'): t = t.split('\n', 1)[-1].rsplit('```', 1)[0]
        items = _json.loads(t)
        days = [AutoFillDay(day_index=i, day_name=DAY_NAMES[i]) for i in range(5)]
        if isinstance(items, list):
            for item in items:
                if not isinstance(item, dict): continue
                ds = str(item.get('date','')).strip()
                period = str(item.get('period','')).strip()
                ts = str(item.get('time_start', item.get('time',''))).strip()
                te = str(item.get('time_end','')).strip()
                act = str(item.get('activity','')).strip()
                res = str(item.get('result','')).strip()
                if not act: continue
                if ts and len(ts)==4 and ':' not in ts: ts = ts[:2]+':'+ts[2:]
                if te and len(te)==4 and ':' not in te: te = te[:2]+':'+te[2:]
                di = -1
                for i in range(5):
                    d = s + timedelta(days=i)
                    if ds in (f"{d.month}月{d.day}日", f"{d.month}.{d.day}"): di = i; break
                if di < 0: continue
                days[di].activities.append(AutoFillActivity(period=period, time_start=ts, time_end=te, activity=act, result=res))
        for day in days:
            if day.activities:
                day.activities.sort(key=lambda a: (a.period!='上午', a.time_start or ''))
        merged = []
        for i in range(5):
            llm_filled = [a for a in days[i].activities if a.activity]
            li = 0
            for a in req.activities:
                if a.day_index != i: continue
                if a.activity:
                    merged.append(AutoFillActivity(period=a.period, time_start=a.time_start, time_end=a.time_end, activity=a.activity, result=a.result))
                else:
                    if li < len(llm_filled):
                        src = llm_filled[li]; li += 1
                        merged.append(AutoFillActivity(period=a.period or src.period, time_start=a.time_start or src.time_start, time_end=a.time_end or src.time_end, activity=src.activity, result=src.result))
                    else:
                        merged.append(AutoFillActivity(period=a.period, time_start=a.time_start, time_end=a.time_end, activity='', result=''))
            days[i].activities = merged
        final_days = [AutoFillDay(day_index=i, day_name=DAY_NAMES[i], activities=days[i].activities) for i in range(5)]
        return AutoFillResponse(days=final_days, warnings=[], detected_dates=[])
    except (Exception, _json.JSONDecodeError) as exc:
        logger.warning("Enrich failed: %s", exc)
        return AutoFillResponse(days=[AutoFillDay(day_index=i, day_name=DAY_NAMES[i], activities=[
            AutoFillActivity(period=a.period, time_start=a.time_start, time_end=a.time_end, activity=a.activity, result=a.result)
            for a in req.activities if a.day_index == i
        ]) for i in range(5)], warnings=[f"完善失败：{exc}"], detected_dates=[])

# ═══════════════ Generate ═══════════════

@router.post("/generate", response_model=GenerateReportResponse)
async def generate_report(req: GenerateReportRequest):
    try:
        s = datetime.strptime(req.start_date, "%Y-%m-%d")
        e = datetime.strptime(req.end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式无效")
    if not req.activities:
        raise HTTPException(status_code=400, detail="请先自动填报或手动添加活动")

    activities_by_day: dict[int, list[dict]] = {}
    for a in req.activities:
        activities_by_day.setdefault(a.day_index, []).append({
            "period": a.period, "time_start": a.time_start, "time_end": a.time_end,
            "activity": a.activity, "result": a.result,
        })

    file_path = None
    if req.template_filename:
        record = get_record(req.template_filename)
        if record:
            for p in record.get("file_paths", []):
                if os.path.isfile(p): file_path = p; break
    if not file_path:
        file_path = excel_ops.find_latest_report(before_date=req.start_date)
    if not file_path:
        raise HTTPException(status_code=404, detail="未找到模板。请先生成至少一份周报。")

    try:
        # ── Use weekly report skill's pipeline (same quality) ──
        sheet = f"{s.month}.{s.day}-{e.month}.{e.day}"
        base = f"26年周工作总结和下周计划-ZB-{s.month}.{s.day}-{e.month}.{e.day}"
        v = 0; fn = f"{base}.xlsx"
        while os.path.exists(os.path.join(OUTPUT_DIR, fn)): v += 1; fn = f"{base}_v{v}.xlsx"
        os.makedirs(OUTPUT_DIR, exist_ok=True); os.makedirs(_DOWNLOAD_DIR, exist_ok=True)
        fp = os.path.join(OUTPUT_DIR, fn); dp = os.path.join(_DOWNLOAD_DIR, fn)

        # 1. Copy template sheet
        wb = openpyxl.load_workbook(file_path)
        src_sheet = excel_ops.get_latest_sheet_name(wb)
        if not src_sheet:
            src_sheet = wb.sheetnames[-1]
        if sheet in wb.sheetnames:
            del wb[sheet]
        excel_ops.copy_sheet(wb, src_sheet, sheet)
        excel_ops.remove_old_data(wb, sheet)

        # 2. Write user's auto-filled content DIRECTLY to the correct D/E cells
        ws = wb[sheet]
        # Get day map first to know row ranges
        day_map_test = excel_ops.get_day_row_map(ws)
        # Build day name → row range for user content placement
        for di in range(5):
            acts = activities_by_day.get(di, [])
            if not acts: continue
            if di not in day_map_test: continue
            first_row, last_row = day_map_test[di]
            # Sort user activities: morning first, then by time
            sorted_acts = sorted(acts, key=lambda a: (
                0 if a.get('period') == '上午' else 1,
                a.get('time_start', '') or ''
            ))
            # Write user content starting from first_row, matching morning/afternoon slots
            morning_written = 0
            afternoon_written = 0
            for a in sorted_acts:
                period = a.get('period', '')
                act = a.get('activity', '') or ''
                res = a.get('result', '') or ''
                # Find the matching row: first available row for this period
                target_row = None
                for r in range(first_row, last_row + 1):
                    cell_period = str(ws.cell(r, 2).value or '').strip()
                    if period == cell_period or (period == '上午' and '上午' in cell_period) or (period == '下午' and '下午' in cell_period):
                        # Check if this row is still empty in D column
                        existing = str(ws.cell(r, 4).value or '').strip()
                        if not existing:
                            target_row = r
                            break
                if target_row:
                    ws.cell(target_row, 4).value = act
                    ws.cell(target_row, 5).value = res if res else '完成'

        # 3. Fill any still-empty cells with simple defaults
        day_map = day_map_test
        for di, (first_row, last_row) in day_map.items():
            for r in range(first_row, last_row + 1):
                d_val = str(ws.cell(r, 4).value or '').strip()
                e_val = str(ws.cell(r, 5).value or '').strip()
                if not d_val:
                    ws.cell(r, 4).value = '日常工作处理'
                if not e_val:
                    ws.cell(r, 5).value = '完成'

        # 4. LLM generate summary (B/C columns only)
        from app.skills.weekly_report.llm_ops import generate_summary as skill_generate_summary
        user_input = ""
        for di in range(5):
            acts = activities_by_day.get(di, [])
            if acts:
                segs = []
                for a in acts:
                    period = a.get('period', '')
                    act = a.get('activity', '') or ''
                    res = a.get('result', '') or ''
                    segs.append(f"{period}：{act}{'（'+res+'）' if res else ''}")
                user_input += f"{DAY_NAMES[di]}：{'；'.join(segs)}\n"
        summary_data = await skill_generate_summary(
            ws, day_map, req.start_date, req.end_date, user_input, [],
            excel_ops.find_summary_row,
        )
        excel_ops.apply_summary_data(ws, summary_data)

        wb.save(fp); wb.close()
        shutil.copy2(fp, dp)

        add_record(fn, sheet, f"{s.month}.{s.day}-{e.month}.{e.day}", [fp, dp])
        return GenerateReportResponse(success=True, message="生成成功", filename=fn, download_url=f"/api/skills/download/{fn}", path=fp, cells_updated=sum(len(a) for a in activities_by_day.values())*2, sheet_name=sheet)
    except HTTPException: raise
    except Exception as exc:
        logger.error("Generate failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成失败：{exc}")

# ═══════════════ Preview ═══════════════

@router.get("/preview/{filename:path}", response_model=ReportPreview)
async def preview_report(filename: str):
    import urllib.parse; filename = urllib.parse.unquote(filename)
    fp = os.path.join(_DOWNLOAD_DIR, filename)
    if not os.path.isfile(fp): fp = os.path.join(OUTPUT_DIR, filename)
    if not os.path.isfile(fp):
        alt = os.path.join(_DATA_OUTPUT_DIR, filename)
        if os.path.isfile(alt): fp = alt
        else: raise HTTPException(status_code=404, detail=f"文件未找到: {filename}")
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
        ts = None
        for name in wb.sheetnames:
            if _re.match(r'^\d+\.\d+-\d+\.\d+$', name): ts = name  # take last (most recent)
        if not ts: ts = wb.sheetnames[-1]
        ws = wb[ts]; mr = ws.max_row
        days, cd, last_period = [], None, ""
        for row in range(2, mr+1):
            a_raw = ws.cell(row,1).value; b_raw = ws.cell(row,2).value; c_raw = ws.cell(row,3).value
            d_raw = ws.cell(row,4).value; e_raw = ws.cell(row,5).value
            # Handle datetime objects
            if hasattr(a_raw, 'strftime'): av = a_raw.strftime("%Y-%m-%d")
            else: av = str(a_raw or "").strip()
            bv = str(b_raw or "").strip(); cv = str(c_raw or "").strip()
            dv = str(d_raw or "").strip(); ev = str(e_raw or "").strip()
            if "本周计划" in av or "本周目标" in av: break
            if bv == "上午" and (not cd or last_period == "下午"):
                idx = len(days)
                ds = av if av else ""
                if idx < 5: cd = {"day_name": DAY_NAMES[idx], "date_str": ds, "rows": []}; days.append(cd)
            if bv: last_period = bv
            period = last_period or ""; clock = cv if cv else ""
            tslot = f"{period} {clock}".strip() if (period and clock) else (period or clock or "")
            if cd and (dv or ev):
                cd["rows"].append({"time_slot": tslot, "plan": dv, "summary": ev})
        sb = sc = ""
        for row in range(2, mr+1):
            av = str(ws.cell(row,1).value or "").strip()
            if "本周计划" in av:
                srr = row + 1
                if srr <= mr: sb = str(ws.cell(srr,2).value or "").strip(); sc = str(ws.cell(srr,3).value or "").strip()
                break
        wb.close()
        return ReportPreview(filename=filename, sheet_name=ts, date_range=ts, days=days, summary_b=sb, summary_c=sc)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"预览失败：{exc}")
