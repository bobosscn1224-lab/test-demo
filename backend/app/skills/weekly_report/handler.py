"""Weekly Report Skill — multi-turn dialogue to fill weekly work report Excel.

Conversation flow:
  1. User triggers → skill asks which week / parses date range
  2. User chooses: A (day-by-day guided) or B (all-at-once fast mode)
  3. Skill collects work info → generates updated Excel from template
  4. Output: new .xlsx file with this week's data

Modules:
  - constants.py   — hardcoded data, prompts, templates
  - date_utils.py  — date parsing and week calculation
  - excel_ops.py   — Excel read/write/copy/summary operations
  - llm_ops.py     — LLM prompt building, generation, parsing
"""
from __future__ import annotations

import logging
from datetime import datetime, timedelta

from app.skills.base import BaseSkill, SkillContext, SkillResult
from app.core.skill_session import SkillSessionHelper

from . import constants as C
from . import date_utils
from . import excel_ops
from . import llm_ops

logger = logging.getLogger(__name__)

_sessions: dict[str, dict] = {}
_helper = SkillSessionHelper(C.SKILL_NAME, _sessions)


class WeeklyReportSkill(BaseSkill):
    name = "weekly_report"
    description = "复制上周周报sheet，通过对话引导了解本周工作，按天更新周报内容"
    triggers = ["写周报", "周报", "生成周报", "帮我写周报"]
    keywords = ["周报", "weekly", "工作周报", "周工作总结"]

    def can_handle(self, message: str) -> bool:
        msg_lower = message.lower()
        for trigger in self.triggers:
            if trigger.lower() in msg_lower:
                return True
        for kw in self.keywords:
            if kw.lower() in msg_lower:
                return True
        return False

    async def execute(self, context: SkillContext) -> SkillResult:
        msg = context.user_message.strip()
        session_id = context.session_id or "default"

        await _helper.restore()

        # --- Check for control commands ---
        exit_kws = ["重新开始", "重来", "取消", "退出", "返回", "算了", "不写了"]
        if len(msg) <= 10 and any(kw in msg for kw in exit_kws):
            await _helper.delete(session_id)
            return SkillResult(success=True, message="好的，已退出周报生成。有什么其他需要随时找我。")

        # --- Ongoing session routing ---
        ongoing = _sessions.get(session_id)
        if ongoing:
            mode = ongoing.get("mode", "guided")
            if mode == "ask_date":
                return await self._handle_ask_date(ongoing, msg, session_id)
            elif mode == "choose":
                return await self._handle_choose(ongoing, msg, session_id)
            elif mode == "fast_collect":
                return await self._handle_fast_collect(ongoing, msg, session_id)
            else:
                return await self._handle_collection(ongoing, msg, session_id)

        # --- New request ---
        if not date_utils.has_date_specifier(msg):
            await _helper.save(session_id, {"mode": "ask_date"})
            weeks = date_utils.build_weeks_data()
            return SkillResult(
                success=True,
                message="好的，请选择要写**哪一周**的周报：",
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "ask_date", "weeks": weeks},
            )

        start_date, end_date = date_utils.parse_date_range(msg)
        return await self._handle_new_request(msg, start_date, end_date, session_id)

    async def _handle_new_request(self, msg: str, start_date: str, end_date: str, session_id: str) -> SkillResult:
        """Handle a new weekly report request with a parsed date range."""
        file_path = excel_ops.find_latest_report()
        if not file_path:
            return SkillResult(
                success=True,
                message=(
                    f"好的，我来帮你生成 {start_date} 到 {end_date} 的周报。\n\n"
                    "但我没有找到之前的周报文件。请上传一份之前的周报 Excel 文件"
                    "（如 `工作周报/` 目录下的文件），我会按一致的格式来生成。"
                ),
                data={"awaiting_upload": True, "skill": "weekly_report", "date_range": f"{start_date}-{end_date}"},
            )

        guided_keywords = ["一天一天", "逐天", "每天聊", "逐日", "一天天"]
        fast_keywords = ["一次性", "一起", "直接生成", "一口气"]
        want_guided = any(kw in msg for kw in guided_keywords)
        want_fast = any(kw in msg for kw in fast_keywords)
        work_details = date_utils.extract_work_details(msg, self.triggers)

        if want_guided:
            return await self._start_guided(file_path, start_date, end_date, session_id)

        if want_fast or work_details:
            if work_details:
                return await self._generate_full(file_path, start_date, end_date, work_details, session_id)
            return SkillResult(
                success=True,
                message=(
                    f"好的，一次性生成 {start_date} 到 {end_date} 的周报。\n\n"
                    "请把这周的重点工作一次性告诉我，比如：\n"
                    "• 周一：上午做了xx，下午做了xx\n• 周二：上午做了xx，下午做了xx\n• ...\n\n"
                    "我会帮你整理成专业表述填入周报。"
                ),
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "fast"},
            )

        # No mode selected — ask A/B
        msg_lower = msg.lower().strip()
        if msg_lower in ("a", "b", "选a", "选b", "a.", "b."):
            if msg_lower.startswith("a") or msg_lower == "a" or msg_lower == "a.":
                return await self._start_guided(file_path, start_date, end_date, session_id)
            else:
                await _helper.save(session_id, {
                    "mode": "fast_collect", "file_path": file_path,
                    "start_date": start_date, "end_date": end_date,
                })
                return SkillResult(
                    success=True,
                    message=(
                        f"好的，一次性生成 {start_date} 到 {end_date} 的周报。\n\n"
                        "请把这周的重点工作一次性告诉我，比如：\n"
                        "• 周一：上午做了xx，下午做了xx\n• 周二：上午做了xx，下午做了xx\n• ...\n\n"
                        "我会帮你整理成专业表述填入周报。"
                    ),
                    data={"skill": "weekly_report", "awaiting_input": True, "mode": "fast"},
                )

        await _helper.save(session_id, {
            "mode": "choose", "file_path": file_path,
            "start_date": start_date, "end_date": end_date,
        })
        s = datetime.strptime(start_date, "%Y-%m-%d")
        return SkillResult(
            success=True,
            message=(
                f"好的，我来帮你生成 {start_date}（周一）到 {end_date}（周日）的周报。\n\n"
                "你想怎么聊？\n"
                "**A. 一天一天聊** — 我每天问你做了什么，逐天收集\n"
                "**B. 一次性聊完** — 你把一周工作一起告诉我，我直接生成\n\n"
                "回复 **A** 或 **B** 即可。"
            ),
            data={"skill": "weekly_report", "awaiting_input": True, "mode": "choose"},
        )

    # ── ask date handler ─────────────────────────────────────────────

    async def _handle_ask_date(self, session: dict, msg: str, session_id: str) -> SkillResult:
        msg_lower = msg.lower().strip()
        if msg_lower in ("a", "b", "选a", "选b", "a.", "b."):
            weeks = date_utils.build_weeks_data()
            return SkillResult(
                success=True,
                message="还没告诉我日期呢～请选择你想写**哪一周**的周报：",
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "ask_date", "weeks": weeks},
            )

        if not date_utils.has_date_specifier(msg):
            weeks = date_utils.build_weeks_data()
            return SkillResult(
                success=True,
                message="请选择你想写**哪一周**的周报，或直接输入日期范围（如 `5月25日-5月31日`）：",
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "ask_date", "weeks": weeks},
            )

        start_date, end_date = date_utils.parse_date_range(msg)

        file_path = excel_ops.find_latest_report()
        if not file_path:
            _sessions.pop(session_id, None)
            return SkillResult(
                success=True,
                message=(
                    f"好的，{start_date} 到 {end_date} 的周报。\n\n"
                    "但我没有找到之前的周报模板文件。请上传一份之前的周报 Excel 文件。"
                ),
                data={"awaiting_upload": True, "skill": "weekly_report", "date_range": f"{start_date}-{end_date}"},
            )

        await _helper.save(session_id, {
            "mode": "choose", "file_path": file_path,
            "start_date": start_date, "end_date": end_date,
        })
        return SkillResult(
            success=True,
            message=(
                f"好的，我来帮你生成 {start_date}（周一）到 {end_date}（周日）的周报。\n\n"
                "你想怎么聊？\n"
                "**A. 一天一天聊** — 我每天问你做了什么\n"
                "**B. 一次性聊完** — 你把一周工作一起告诉我\n\n"
                "回复 **A** 或 **B** 即可。"
            ),
            data={"skill": "weekly_report", "awaiting_input": True, "mode": "choose"},
        )

    # ── mode choice / fast collect ────────────────────────────────────

    async def _handle_choose(self, session: dict, msg: str, session_id: str) -> SkillResult:
        msg_lower = msg.lower().strip()
        file_path = session["file_path"]
        start_date = session["start_date"]
        end_date = session["end_date"]

        if msg_lower.startswith("a") or msg_lower in ("a", "a.", "选a"):
            _sessions.pop(session_id, None)
            return await self._start_guided(file_path, start_date, end_date, session_id)
        elif msg_lower.startswith("b") or msg_lower in ("b", "b.", "选b"):
            session["mode"] = "fast_collect"
            return SkillResult(
                success=True,
                message=(
                    f"好的，一次性生成 {start_date} 到 {end_date} 的周报。\n\n"
                    "请把这周的重点工作一次性告诉我，比如：\n"
                    "• 周一：上午做了xx，下午做了xx\n• 周二：上午做了xx，下午做了xx\n• ...\n\n"
                    "我会帮你整理成专业表述填入周报。"
                ),
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "fast"},
            )
        return SkillResult(
            success=True,
            message="请回复 **A**（一天一天聊）或 **B**（一次性聊完），我来帮你开始。",
            data={"skill": "weekly_report", "awaiting_input": True, "mode": "choose"},
        )

    async def _handle_fast_collect(self, session: dict, msg: str, session_id: str) -> SkillResult:
        work_details = date_utils.extract_work_details(msg, self.triggers)
        if not work_details:
            return SkillResult(
                success=True,
                message="请描述这周的重点工作，内容太少我无法生成。可以按天列出，比如：周一上午xx、下午xx...",
                data={"skill": "weekly_report", "awaiting_input": True, "mode": "fast"},
            )
        _sessions.pop(session_id, None)
        return await self._generate_full(session["file_path"], session["start_date"], session["end_date"], work_details, session_id)

    # ── guided day-by-day collection ──────────────────────────────────

    async def _start_guided(self, file_path: str, start_date: str, end_date: str, session_id: str) -> SkillResult:
        try:
            last_sheet_content = excel_ops.read_sheet_content(file_path, excel_ops.get_latest_sheet_name)
        except Exception:
            last_sheet_content = "无法读取上周内容，将使用默认模板格式。"

        await _helper.save(session_id, {
            "file_path": file_path, "start_date": start_date, "end_date": end_date,
            "day_index": 0, "collected": {}, "last_sheet_content": last_sheet_content,
        })

        s = datetime.strptime(start_date, "%Y-%m-%d")
        return SkillResult(
            success=True,
            message=f"好的，我们一天一天来。从 **周一（{s.month}月{s.day}日）** 开始——\n周一主要做了哪些工作？（可以简单描述上午和下午的内容）",
            data={"skill": "weekly_report", "awaiting_input": True, "day": "周一"},
        )

    async def _handle_collection(self, session: dict, msg: str, session_id: str) -> SkillResult:
        day_idx = session["day_index"]
        day_name = C.DAY_NAMES[day_idx]

        session["collected"][day_idx] = msg
        next_idx = day_idx + 1

        if next_idx >= 5:
            result = await self._generate_from_collected(session, session_id)
            _sessions.pop(session_id, None)
            return result

        session["day_index"] = next_idx
        next_name = C.DAY_NAMES[next_idx]
        start_date = datetime.strptime(session["start_date"], "%Y-%m-%d")
        next_date = start_date + timedelta(days=next_idx)

        collected_summary = self._build_collected_summary(session)

        return SkillResult(
            success=True,
            message=(
                f"收到，已记录{day_name}的工作内容。\n\n"
                f"{collected_summary}\n"
                f"接下来，**{next_name}（{next_date.month}月{next_date.day}日）** 主要做了什么？"
            ),
            data={"skill": "weekly_report", "awaiting_input": True, "day": next_name, "progress": f"{next_idx}/5"},
        )

    def _build_collected_summary(self, session: dict) -> str:
        lines = ["📋 已收集："]
        for i in sorted(session["collected"].keys()):
            content = session["collected"][i][:80]
            lines.append(f"  {C.DAY_NAMES[i]}：{content}...")
        return "\n".join(lines)

    # ── full generation ──────────────────────────────────────────────

    async def _generate_from_collected(self, session: dict, session_id: str) -> SkillResult:
        start_date = session["start_date"]
        end_date = session["end_date"]
        file_path = session["file_path"]

        user_input_parts = []
        for i in range(5):
            if i in session["collected"]:
                user_input_parts.append(f"{C.DAY_NAMES[i]}：{session['collected'][i]}")
        user_input = "\n".join(user_input_parts)

        return await self._generate_full(file_path, start_date, end_date, user_input, session_id)

    async def _generate_full(
        self, file_path: str, start_date: str, end_date: str, user_input: str, session_id: str
    ) -> SkillResult:
        """Full generation: copy template sheet, update content, save."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path)
            latest_sheet_name = excel_ops.get_latest_sheet_name(wb)

            s = datetime.strptime(start_date, "%Y-%m-%d")
            e = datetime.strptime(end_date, "%Y-%m-%d")
            new_sheet_name = f"{s.month}.{s.day}-{e.month}.{e.day}"

            if new_sheet_name in wb.sheetnames:
                wb.close()
                return SkillResult(success=False, message=f"工作表 '{new_sheet_name}' 已存在。")

            if latest_sheet_name:
                excel_ops.copy_sheet(wb, latest_sheet_name, new_sheet_name)
            else:
                ws = wb.create_sheet(title=new_sheet_name)
                ws["A1"] = "日期/星期"
                ws["B1"] = "时间"
                ws["C1"] = "时间"
                ws["D1"] = "本周计划"
                ws["E1"] = "本周总结（完成情况及改进，有什么问题）"

            excel_ops.remove_old_data(wb, new_sheet_name)

            kb_patterns = await llm_ops.search_knowledge_base()

            fill_data, day_map = await llm_ops.generate_updated_content(
                wb, new_sheet_name, start_date, end_date, user_input, kb_patterns,
                excel_ops.get_day_row_map, excel_ops.read_existing_structure,
            )

            ws = wb[new_sheet_name]
            excel_ops.merge_date_cells(ws, day_map)

            max_data_row = max(r[1] for r in day_map.values()) if day_map else ws.max_row
            excel_ops.apply_fill_data(ws, fill_data, max_data_row)
            excel_ops.fill_empty_cells(ws, day_map, start_date)

            summary_data = await llm_ops.generate_summary(
                ws, day_map, start_date, end_date, user_input, kb_patterns,
                excel_ops.find_summary_row,
            )
            excel_ops.apply_summary_data(ws, summary_data)

            # Save to output directory
            import os as _os
            _os.makedirs(C.OUTPUT_DIR, exist_ok=True)
            new_filename = f"26年周工作总结和下周计划-ZB-{s.month}.{s.day}-{e.month}.{e.day}.xlsx"
            new_filepath = _os.path.join(C.OUTPUT_DIR, new_filename)
            wb.save(new_filepath)
            wb.close()

            # Copy to unified outputs dir for download
            from app.services._paths import OUTPUTS_DIR
            data_output = str(OUTPUTS_DIR)
            _os.makedirs(data_output, exist_ok=True)
            wb2 = openpyxl.load_workbook(new_filepath)
            wb2.save(_os.path.join(data_output, new_filename))
            wb2.close()

            summary = self._build_summary(fill_data, start_date, end_date, new_filename, new_sheet_name)

            return SkillResult(
                success=True,
                message=summary,
                data={
                    "download_url": f"/api/skills/download/{new_filename}",
                    "filename": new_filename,
                    "path": new_filepath,
                },
                follow_up_action="download",
            )
        except Exception as e:
            logger.error("Weekly report generation failed: %s", e, exc_info=True)
            return SkillResult(success=False, message=f"周报生成过程出错：{e}")

    def _build_summary(self, fill_data: list[dict], start_date: str, end_date: str, filename: str, sheet_name: str) -> str:
        days = set()
        for item in fill_data:
            if isinstance(item, dict) and item.get("format") == "date":
                days.add(str(item.get("value", ""))[:10])

        cells_updated = len([i for i in fill_data if isinstance(i, dict) and i.get("format") != "date"])
        lines = [
            "✅ 周报已生成！",
            "",
            f"📅 时间范围: {start_date} ~ {end_date}",
            f"📋 新增工作表: {sheet_name}",
            f"📄 文件: {filename}",
            f"📊 覆盖 {len(days)} 天，更新 {cells_updated} 个单元格",
            "",
            "📁 工作流说明：",
            "  ① 已从上周工作表复制内容作为模板",
            f"  ② 新建工作表「{sheet_name}」",
            "  ③ 已更新D列（本周计划）和E列（本周总结）",
            "  ④ 已修正A列日期为本周日期",
        ]

        preview_items = [i for i in fill_data if isinstance(i, dict) and i.get("format") != "date"][:6]
        if preview_items:
            lines.append("")
            lines.append("📝 内容预览:")
            for item in preview_items:
                lines.append(f"  • {item.get('cell', '?')}: {str(item.get('value', ''))[:60]}")

        return "\n".join(lines)
