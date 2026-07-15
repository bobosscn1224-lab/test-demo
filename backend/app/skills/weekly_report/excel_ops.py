"""Weekly report Excel operations — read, copy, write, merge."""

import copy
import os
import re
from datetime import datetime, timedelta

from .constants import (
    DAY_NAMES, OUTPUT_DIR, TEMPLATE_FILE,
    MORNING_VARIANTS_D, MORNING_VARIANTS_E,
    AFTERNOON_VARIANTS_D, AFTERNOON_VARIANTS_E,
    DEFAULT_D_LAST, DEFAULT_E_LAST,
)


def find_latest_report(before_date: str | None = None) -> str | None:
    """Find the best template to use as base.

    Priority:
      1. If before_date given: find the latest report whose end_date < before_date
      2. Fallback: use the fixed TEMPLATE_FILE if it exists
      3. Last resort: return None
    """
    import glob as _glob
    import re as _re
    from datetime import datetime as _dt

    # Collect all existing reports with their date ranges
    candidates: list[tuple[str, _dt]] = []
    for d in [OUTPUT_DIR, os.path.dirname(TEMPLATE_FILE)]:
        if not os.path.isdir(d):
            continue
        for fpath in _glob.glob(os.path.join(d, "*.xlsx")):
            fname = os.path.basename(fpath)
            if fname.startswith("~$"):
                continue
            m = _re.search(r'(\d+)\.(\d+)-(\d+)\.(\d+)', fname)
            if not m:
                continue
            m1, d1, m2, d2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
            try:
                end_date = _dt(_dt.now().year, m2, d2)
                candidates.append((fpath, end_date))
            except ValueError:
                continue

    if not candidates:
        # Fallback to fixed template
        return TEMPLATE_FILE if os.path.isfile(TEMPLATE_FILE) else None

    # Sort by end_date descending
    candidates.sort(key=lambda x: x[1], reverse=True)

    if before_date:
        try:
            bd = _dt.strptime(before_date, "%Y-%m-%d")
            # Find the latest report that ended before the target week start
            for fpath, end_date in candidates:
                if end_date < bd:
                    return fpath
        except ValueError:
            pass

    # No before_date filter or no match — return the most recent
    return candidates[0][0] if candidates else (
        TEMPLATE_FILE if os.path.isfile(TEMPLATE_FILE) else None
    )


def get_latest_sheet_name(wb) -> str | None:
    """Get the most recent date-range sheet (e.g. '5.11-5.15')."""
    date_range_sheets = []
    for s in wb.sheetnames:
        if re.match(r'^\d+\.\d+-\d+\.\d+$', s):
            m = re.match(r'^(\d+)\.(\d+)-', s)
            if m:
                month, day = int(m.group(1)), int(m.group(2))
                date_range_sheets.append(((month, day), s))
    if date_range_sheets:
        date_range_sheets.sort()
        return date_range_sheets[-1][1]
    skip = {"Sheet1", "sheet1"}
    candidates = [s for s in wb.sheetnames if s not in skip and re.match(r'^\d', s)]
    if candidates:
        return candidates[-1]
    return None


def copy_sheet(wb, source_name: str, target_name: str):
    """Copy a worksheet within the same workbook with formatting."""
    from openpyxl.utils import get_column_letter

    ws_source = wb[source_name]
    ws_target = wb.create_sheet(title=target_name)

    for row in ws_source.iter_rows(min_row=1, max_row=ws_source.max_row, max_col=ws_source.max_column):
        for cell in row:
            new_cell = ws_target.cell(row=cell.row, column=cell.column)
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy.copy(cell.alignment)

    for merged_range in ws_source.merged_cells.ranges:
        ws_target.merge_cells(str(merged_range))

    for col_idx in range(1, ws_source.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_source.column_dimensions:
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width

    for row_idx in range(1, ws_source.max_row + 1):
        if row_idx in ws_source.row_dimensions:
            ws_target.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height


def remove_old_data(wb, current_sheet: str):
    """Remove 2025-dated sheets and clear 2025 rows from current sheet."""
    sheets_to_remove = []
    for name in wb.sheetnames:
        if name == current_sheet:
            continue
        if '2025' in name:
            sheets_to_remove.append(name)

    for name in sheets_to_remove:
        try:
            del wb[name]
        except Exception:
            pass

    ws = wb[current_sheet]
    rows_to_clear = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and '2025' in str(cell.value):
            rows_to_clear.append(cell.row)

    for r in rows_to_clear:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=r, column=col).value = None


def read_sheet_content(file_path: str, get_sheet_fn) -> str:
    """Read the latest sheet content as a text summary for context."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    sheet_name = get_sheet_fn(wb)
    if not sheet_name:
        wb.close()
        return ""

    ws = wb[sheet_name]
    lines = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 35), values_only=True):
        parts = [str(c) if c is not None else "" for c in row[:5]]
        if any(parts):
            lines.append(" | ".join(parts))

    wb.close()
    return "\n".join(lines)


def read_existing_structure(ws) -> str:
    """Read first 20 rows as format sample."""
    lines = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        parts = [str(c).strip() if c is not None else "" for c in row[:5]]
        if any(parts):
            lines.append(" | ".join(parts))
    lines.append(f"... (总行数: {ws.max_row})")
    return "\n".join(lines)


def get_day_row_map(ws) -> dict[int, tuple[int, int]]:
    """Parse which rows belong to each day (Mon=0..Fri=4)."""
    # Unmerge all column-A merged ranges first
    merged_to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col == 1:
            merged_to_unmerge.append(str(merged_range))
    for mr_str in merged_to_unmerge:
        try:
            ws.unmerge_cells(mr_str)
        except Exception:
            pass

    # Clear old DATE values from column A
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val is not None:
            val_str = str(val).strip()
            if val_str and re.match(r'^\d{4}-\d{2}-\d{2}', val_str):
                ws.cell(row, 1).value = None

    # Find B-column "上午" cells which mark day starts
    day_boundaries = [2]
    for row in range(3, ws.max_row + 1):
        b_val = str(ws.cell(row, 2).value or "").strip()
        if b_val == "上午":
            day_boundaries.append(row)

    day_boundaries = day_boundaries[:5]

    # Build day map
    day_map = {}
    for i, start_row in enumerate(day_boundaries):
        if i + 1 < len(day_boundaries):
            end_row = day_boundaries[i + 1] - 1
        else:
            end_row = ws.max_row
            for r in range(ws.max_row, start_row, -1):
                a_val = str(ws.cell(r, 1).value or "")
                d_val = str(ws.cell(r, 4).value or "")
                combined = a_val + d_val
                if any(kw in combined for kw in ["重点", "目标", "计划", "课题", "下周"]):
                    end_row = r - 1
                    break
        if end_row >= start_row:
            day_map[i] = (start_row, end_row)

    return day_map


def merge_date_cells(ws, day_map: dict):
    """Re-merge A-column cells for each day's row range after writing dates."""
    for day_idx, (first_row, last_row) in day_map.items():
        if last_row > first_row:
            try:
                ws.merge_cells(None, first_row, 1, last_row, 1)
            except Exception:
                pass


def find_summary_row(ws, day_map: dict) -> int:
    """Find the row for B/C summary fill."""
    max_data_row = max(r[1] for r in day_map.values()) if day_map else ws.max_row
    for r in range(max_data_row + 1, ws.max_row + 1):
        a_val = str(ws.cell(r, 1).value or "").strip()
        if "本周计划" in a_val or "本周目" in a_val:
            content_row = r + 1
            if content_row <= ws.max_row:
                return content_row
            return r
    for r in range(max_data_row + 1, ws.max_row + 1):
        a_val = str(ws.cell(r, 1).value or "").strip()
        if "计划" in a_val or "目标" in a_val:
            continue
        return r
    return max_data_row + 2


def apply_fill_data(ws, fill_data: list[dict], max_data_row: int | None = None):
    """Apply LLM-generated fill data to D/E columns."""
    clear_end = max_data_row if max_data_row is not None else ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=clear_end, min_col=4, max_col=5):
        for cell in row:
            cell.value = None

    for item in fill_data:
        if not (isinstance(item, dict) and "cell" in item):
            continue
        try:
            value = item.get("value", "")
            cell_format = item.get("format", "")
            cell_ref = item["cell"]
            m = re.match(r'([A-Z]+)(\d+)', cell_ref)
            if m:
                row_num = int(m.group(2))
                if row_num > clear_end or row_num < 2:
                    continue
            if cell_format == "date" and isinstance(value, str):
                try:
                    value = datetime.strptime(value, "%Y-%m-%d")
                except ValueError:
                    pass
            ws[cell_ref] = value
        except Exception:
            pass


def fill_empty_cells(ws, day_map: dict, start_date: str):
    """Fill remaining empty D/E cells within work hours. Skip out-of-hours rows."""
    for day_idx, (first_row, last_row) in day_map.items():
        for r in range(first_row, last_row + 1):
            b_val = str(ws.cell(r, 2).value or "").strip()
            c_val = str(ws.cell(r, 3).value or "").strip()
            d_val = str(ws.cell(r, 4).value or "").strip()
            e_val = str(ws.cell(r, 5).value or "").strip()
            # Skip rows outside work hours
            if b_val == '上午' and c_val and c_val < '09:00': continue
            if b_val == '下午' and c_val and c_val >= '18:00': continue
            if not d_val: ws.cell(r, 4).value = "—"
            if not e_val: ws.cell(r, 5).value = "—"



def apply_summary_data(ws, summary_data: list[dict]):
    """Apply summary data to B/C columns only."""
    for item in summary_data:
        if not (isinstance(item, dict) and "cell" in item):
            continue
        try:
            cell_ref = item["cell"]
            col_letter = cell_ref[0] if cell_ref else ""
            if col_letter not in ("B", "C"):
                continue
            value = item.get("value", "")
            m = re.match(r'([A-Z]+)(\d+)', cell_ref)
            if m and int(m.group(2)) > ws.max_row:
                continue
            if isinstance(value, str) and value.strip():
                value = value.strip()
            ws[cell_ref] = value
        except Exception:
            pass
