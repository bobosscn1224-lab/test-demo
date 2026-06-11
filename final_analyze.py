# -*- coding: utf-8 -*-
"""Final analysis: classify meeting themes based on ALL column content.

Strategy: use indicator columns (15-20) as primary signal, then text analysis
in other columns (4, 8, 9, 10, 11, 12) as supplementary signal for MISSING themes.
"""
import openpyxl
import os

SRC = r"D:\数字分身\本地知识库\项目分析会汇总数据表_合并1.xlsx"
DST = r"D:\数字分身\项目分析会汇总数据表_更新调整.xlsx"

# ── Theme detection rules ──
# Primary: indicator columns with data
#   - Col 15 (解决方案评分), Col 16 (价值主张) → 解决方案设计与评审
#   - Col 17 (招标策略-下一步计划), Col 18 (招标策略-折扣) → 招标策略制定
#   - Col 19 (投标策略-下一步行动计划), Col 20 (投标策略-折扣) → 投标策略制定
#
# Supplementary: text keyword matching in content-heavy columns (4, 8, 9, 10, 11)
#   These help find MISSING themes when indicator columns are empty

# Precise keywords for each theme (avoid overly broad terms)
SOLUTION_KW = [
    "解决方案设计", "方案设计与评审", "方案评审", "技术评审",
    "价值主张", "解决方案评分", "技术方案", "方案设计",
    "产品选型", "POC测试", "测试方案", "方案评估",
    "技术交流", "方案汇报", "方案演示", "方案讲解",
    "技术建议书", "方案建议书", "架构设计", "配置方案",
    "技术验证", "功能测试", "性能测试", "技术评估",
]

BID_KW = [
    "投标策略", "投标方案", "投标报价", "投标评审",
    "投标决策", "投标分析", "投标策划", "投标组织",
    "投标准备", "标书制作", "标书评审", "投标文件",
    "讲标", "述标", "控标", "投标成本",
    "投标风险", "投标授权", "投标保证金", "投标答疑",
    "中标", "竞标分析", "投标计划", "标前会",
    "投标价格", "投标折扣", "报价策略",
]

TENDER_KW = [
    "招标策略", "招标文件", "招标要求", "招标参数",
    "招标清单", "招标评分", "招标答疑", "招标公告",
    "招标方式", "公开招标", "邀请招标", "竞争性谈判",
    "竞争性磋商", "单一来源", "询价采购", "框架协议",
    "供应商选择", "供应商评估", "供应商入围", "资格审查",
    "采购流程", "采购策略", "寻源",
]

def has_data(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return len(s) > 0 and s not in ("无", "N/A", "-", "暂无", "None", "nan")

def get_text(val) -> str:
    if val is None:
        return ""
    return str(val).strip()

def detect_themes(row_cells, row_idx):
    """Detect ALL themes for a row. Returns (themes_str, reasons_str)."""
    reasons = {}

    # ── Step 1: Check indicator columns ──
    sol_indicators = []
    for col in [15, 16]:
        if has_data(row_cells[col]):
            sol_indicators.append(col)

    tender_indicators = []
    for col in [17, 18]:
        if has_data(row_cells[col]):
            tender_indicators.append(col)

    bid_indicators = []
    for col in [19, 20]:
        if has_data(row_cells[col]):
            bid_indicators.append(col)

    # ── Step 2: Text analysis on content columns ──
    # Focus on: Col 4 (项目分析会名称), Col 8 (分析会参与人),
    #            Col 9 (下一步行动计划), Col 10 (商机阶段), Col 11 (项目名称)
    text_cols = [4, 8, 9, 10, 11, 12]
    all_text = ""
    for col in text_cols:
        if has_data(row_cells[col]):
            all_text += get_text(row_cells[col]) + " "

    # Also include the original theme col for text matching
    all_text += get_text(row_cells[3]) + " "

    # Find keyword matches in text
    sol_text_matches = [kw for kw in SOLUTION_KW if kw in all_text]
    bid_text_matches = [kw for kw in BID_KW if kw in all_text]
    tender_text_matches = [kw for kw in TENDER_KW if kw in all_text]

    # ── Step 3: Combine evidence ──
    # A theme is present if EITHER indicator columns have data OR text has relevant keywords
    themes = []
    reason_parts = []

    # Solution theme
    sol_evidence = []
    if sol_indicators:
        sol_evidence.append(f"指标列Col{sol_indicators}有数据")
    if sol_text_matches:
        sol_evidence.append(f"文本含关键词: {', '.join(sol_text_matches[:3])}")
    if sol_evidence:
        themes.append("解决方案设计与评审")
        reason_parts.append(f"[解决方案设计与评审] {'; '.join(sol_evidence)}")

    # Tender theme
    tender_evidence = []
    if tender_indicators:
        tender_evidence.append(f"指标列Col{tender_indicators}有数据")
    if tender_text_matches:
        tender_evidence.append(f"文本含关键词: {', '.join(tender_text_matches[:3])}")
    if tender_evidence:
        themes.append("招标策略制定")
        reason_parts.append(f"[招标策略制定] {'; '.join(tender_evidence)}")

    # Bid theme
    bid_evidence = []
    if bid_indicators:
        bid_evidence.append(f"指标列Col{bid_indicators}有数据")
    if bid_text_matches:
        bid_evidence.append(f"文本含关键词: {', '.join(bid_text_matches[:3])}")
    if bid_evidence:
        themes.append("投标策略制定")
        reason_parts.append(f"[投标策略制定] {'; '.join(bid_evidence)}")

    if not themes:
        return "未分类", "所有指标列均无数据，且文本中未发现明确主题关键词"

    return "、".join(themes), "；".join(reason_parts)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    # Read all rows into memory for faster processing
    all_rows = []
    for row_idx in range(1, ws.max_row + 1):
        cells = {}
        for col in range(1, ws.max_column + 1):
            cells[col] = ws.cell(row_idx, col).value
        all_rows.append(cells)

    # Add new columns
    theme_col = ws.max_column + 1
    reason_col = ws.max_column + 2
    old_theme_col = ws.max_column + 3  # Keep old theme for comparison
    ws.cell(1, theme_col, "判定主题")
    ws.cell(1, reason_col, "判定理由")
    ws.cell(1, old_theme_col, "原主题(对比)")

    stats = {"解决方案设计与评审": 0, "投标策略制定": 0, "招标策略制定": 0,
             "混合": 0, "未分类": 0}
    changes = []  # rows where new theme differs from old

    for row_idx in range(2, ws.max_row + 1):
        cells = all_rows[row_idx - 1]
        themes, reasons = detect_themes(cells, row_idx)

        old_theme = get_text(cells.get(3, ""))

        ws.cell(row_idx, theme_col, themes)
        ws.cell(row_idx, reason_col, reasons)
        ws.cell(row_idx, old_theme_col, old_theme)

        # Stats
        if "、" in themes:
            stats["混合"] += 1
        elif themes in stats:
            stats[themes] += 1
        else:
            stats["未分类"] += 1

        # Track changes
        old_set = set()
        for t in old_theme.replace("、", ",").replace("+", ",").split(","):
            t = t.strip()
            if t:
                old_set.add(t)

        # Normalize theme names for comparison
        theme_map = {
            "解决方案设计与评审": "解决方案设计与评审",
            "投标策略制定": "投标策略制定",
            "招标策略制定": "招标策略制定",
        }
        new_set = set()
        for t in themes.split("、"):
            t = t.strip()
            if t in theme_map:
                new_set.add(theme_map[t])

        if old_set != new_set:
            name = get_text(cells.get(4, ""))
            changes.append({
                "row": row_idx,
                "name": name[:80],
                "old": old_theme,
                "new": themes,
                "added": new_set - old_set,
                "removed": old_set - new_set,
            })

    wb.save(DST)

    print(f"File saved: {DST}")
    print(f"Total rows: {ws.max_row - 1}")
    print(f"\nStats:")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    print(f"\nRows with changed theme: {len(changes)}")
    print("\n=== Change details ===")
    for c in changes:
        print(f"\nRow {c['row']}: {c['name']}")
        print(f"  Old: {c['old']}")
        print(f"  New: {c['new']}")
        if c['added']:
            print(f"  +Added: {c['added']}")
        if c['removed']:
            print(f"  -Removed: {c['removed']}")

    # Save detailed change log
    log_path = r"D:\数字分身\analysis_changes.txt"
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"Theme Changes Report - {len(changes)} rows changed\n")
        f.write("=" * 60 + "\n\n")
        for c in changes:
            f.write(f"Row {c['row']}: {c['name']}\n")
            f.write(f"  Old: {c['old']}\n")
            f.write(f"  New: {c['new']}\n")
            if c['added']:
                f.write(f"  +Added: {c['added']}\n")
            if c['removed']:
                f.write(f"  -Removed: {c['removed']}\n")
            f.write("\n")
    print(f"\nDetailed log: {log_path}")


if __name__ == "__main__":
    main()
