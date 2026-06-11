# -*- coding: utf-8 -*-
"""
Objective theme analysis:
  - Baseline: Keep original Col 3 theme (expert-labeled, never remove)
  - Add themes based on TWO evidence sources:
    1. Indicator columns (Col 15-20): data in columns that original theme didn't capture
    2. Text keywords in indicator column content: cross-theme discussion signals
  - Only write reasons for rows where themes were ADDED (user will audit)
"""
import openpyxl

SRC = r"D:\数字分身\本地知识库\项目分析会汇总数据表_合并1.xlsx"
DST = r"D:\数字分身\项目分析会汇总数据表_更新调整.xlsx"

# ── Cross-theme keywords (objective, specific) ──
SOLUTION_KW = [
    "方案设计", "方案评审", "技术方案", "技术评审", "产品选型",
    "方案评估", "技术交流", "方案汇报", "方案演示", "POC",
    "测试方案", "技术验证", "方案建议书", "技术建议书",
    "配置方案", "方案讲解", "架构设计", "技术评估",
    "方案优化", "方案确认", "技术参数", "价值主张",
]

BID_KW = [
    "投标策略", "投标方案", "投标报价", "投标评审", "投标决策",
    "投标分析", "投标策划", "投标组织", "标书", "投标文件",
    "讲标", "述标", "控标", "投标成本", "投标风险",
    "投标价格", "竞标分析", "报价策略", "投标计划",
    "投标折扣", "标书制作", "标书评审",
]

TENDER_KW = [
    "招标策略", "招标文件", "招标要求", "招标参数",
    "招标清单", "招标评分", "招标答疑", "招标公告",
    "招标方式", "公开招标", "邀请招标", "竞争性谈判",
    "竞争性磋商", "单一来源", "询价采购", "框架协议",
    "供应商选择", "供应商评估", "供应商入围", "资格审查",
    "采购流程", "采购策略", "寻源",
]

def has_data(val):
    if val is None: return False
    s = str(val).strip()
    return len(s) > 0 and s not in ("无", "N/A", "-", "暂无", "None", "nan")

def t(val):
    if val is None: return ""
    return str(val).strip()

def analyze_row(ws, row):
    old_theme = t(ws.cell(row, 3).value)
    old_set = set()
    for tp in old_theme.replace("、", ",").replace("+", ",").split(","):
        tp = tp.strip()
        if tp:
            old_set.add(tp)

    # ── Evidence 1: Indicator columns ──
    sol_ind = has_data(ws.cell(row, 15).value) or has_data(ws.cell(row, 16).value)
    tender_ind = has_data(ws.cell(row, 17).value) or has_data(ws.cell(row, 18).value)
    bid_ind = has_data(ws.cell(row, 19).value) or has_data(ws.cell(row, 20).value)

    # ── Evidence 2: Text keywords in ALL columns ──
    all_text = ""
    col_kw_map = {}  # kw → [(col, header)]
    for col in range(1, ws.max_column + 1):
        if has_data(ws.cell(row, col).value):
            txt = t(ws.cell(row, col))
            all_text += txt + " "
            header = t(ws.cell(1, col).value)
            for kw in SOLUTION_KW + BID_KW + TENDER_KW:
                if kw in txt:
                    col_kw_map.setdefault(kw, []).append((col, header))

    sol_kws = [kw for kw in SOLUTION_KW if kw in all_text]
    bid_kws = [kw for kw in BID_KW if kw in all_text]
    tender_kws = [kw for kw in TENDER_KW if kw in all_text]

    # ── Combine evidence ──
    # Business linkage (user-confirmed): bid → solution (can't bid without solution)
    # and solution → bid (solutions designed for bidding context)
    has_solution = sol_ind or bool(sol_kws) or bid_ind or bool(bid_kws)
    has_bid = bid_ind or bool(bid_kws) or sol_ind or bool(sol_kws)
    has_tender = tender_ind or bool(tender_kws)

    # Build final set (add to existing, never remove)
    final_set = old_set.copy()
    if has_solution: final_set.add("解决方案设计与评审")
    if has_tender: final_set.add("招标策略制定")
    if has_bid: final_set.add("投标策略制定")

    # ── Build ADDED reasons only ──
    added_reasons = []

    sol_in_old = "解决方案设计与评审" in old_set
    tender_in_old = "招标策略制定" in old_set
    bid_in_old = "投标策略制定" in old_set

    if "解决方案设计与评审" in final_set and not sol_in_old:
        evidence = []
        if sol_ind:
            evidence.append("解决方案评分/价值主张列有数据")
        if sol_kws:
            details = []
            for kw in sol_kws[:4]:
                cols = col_kw_map.get(kw, [])
                col_refs = [f"Col{c}({h})" for c, h in cols[:2]]
                details.append(f"'{kw}'见于{'/'.join(col_refs)}")
            evidence.append("文本关键词: " + "; ".join(details))
        if not sol_ind and not sol_kws:
            # Added due to bid→solution business linkage
            evidence.append("MO流程联动规则：投标策略制定必然涉及解决方案讨论(用户确认)")
        added_reasons.append(f"+解决方案设计与评审 | {'; '.join(evidence)}")

    if "招标策略制定" in final_set and not tender_in_old:
        evidence = []
        if tender_ind:
            evidence.append("招标策略/折扣列有数据")
        if tender_kws:
            details = []
            for kw in tender_kws[:4]:
                cols = col_kw_map.get(kw, [])
                col_refs = [f"Col{c}({h})" for c, h in cols[:2]]
                details.append(f"'{kw}'见于{'/'.join(col_refs)}")
            evidence.append("文本关键词: " + "; ".join(details))
        added_reasons.append(f"+招标策略制定 | {'; '.join(evidence)}")

    if "投标策略制定" in final_set and not bid_in_old:
        evidence = []
        if bid_ind:
            evidence.append("投标策略/折扣列有数据")
        if bid_kws:
            details = []
            for kw in bid_kws[:4]:
                cols = col_kw_map.get(kw, [])
                col_refs = [f"Col{c}({h})" for c, h in cols[:2]]
                details.append(f"'{kw}'见于{'/'.join(col_refs)}")
            evidence.append("文本关键词: " + "; ".join(details))
        if not bid_ind and not bid_kws:
            evidence.append("MO流程联动规则：解决方案设计与评审必然涉及投标策略考量(用户确认)")
        added_reasons.append(f"+投标策略制定 | {'; '.join(evidence)}")

    if not final_set:
        return "未分类", ""

    theme_order = ["解决方案设计与评审", "招标策略制定", "投标策略制定"]
    sorted_themes = [t for t in theme_order if t in final_set]
    new_theme = "、".join(sorted_themes)
    reason = " | ".join(added_reasons) if added_reasons else ""

    return new_theme, reason


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    theme_col = ws.max_column + 1
    reason_col = ws.max_column + 2
    old_theme_col = ws.max_column + 3
    ws.cell(1, theme_col, "判定主题")
    ws.cell(1, reason_col, "补充原因(仅新增主题)")
    ws.cell(1, old_theme_col, "原主题(对比)")

    changes = []
    stats = {}

    for row in range(2, ws.max_row + 1):
        new_theme, reason = analyze_row(ws, row)
        old_theme = t(ws.cell(row, 3).value)

        ws.cell(row, theme_col, new_theme)
        ws.cell(row, reason_col, reason)
        ws.cell(row, old_theme_col, old_theme)

        theme_count = len(new_theme.split("、")) if new_theme != "未分类" else 0
        if theme_count > 1:
            stats["混合(多主题)"] = stats.get("混合(多主题)", 0) + 1
        elif new_theme in stats:
            stats[new_theme] += 1
        else:
            stats[new_theme] = stats.get(new_theme, 0) + 1

        old_set = set()
        for tp in old_theme.replace("、", ",").replace("+", ",").split(","):
            tp = tp.strip()
            if tp:
                old_set.add(tp)
        new_set = set(t for t in new_theme.split("、"))

        if old_set != new_set:
            proj = t(ws.cell(row, 1).value)[:80]
            changes.append({
                "row": row, "project": proj,
                "old": old_theme, "new": new_theme, "reason": reason,
                "added": new_set - old_set, "removed": old_set - new_set,
            })

    wb.save(DST)

    total = ws.max_row - 1
    sol_count = sum(1 for row in range(2, ws.max_row+2) if "解决方案设计与评审" in str(ws.cell(row, theme_col).value or ""))
    bid_count = sum(1 for row in range(2, ws.max_row+2) if "投标策略制定" in str(ws.cell(row, theme_col).value or ""))
    tender_count = sum(1 for row in range(2, ws.max_row+2) if "招标策略制定" in str(ws.cell(row, theme_col).value or ""))

    print(f"File: {DST}")
    print(f"Total: {total} rows")
    print(f"\nPer-theme coverage:")
    print(f"  解决方案设计与评审: {sol_count} ({100*sol_count//total}%)")
    print(f"  招标策略制定: {tender_count} ({100*tender_count//total}%)")
    print(f"  投标策略制定: {bid_count} ({100*bid_count//total}%)")

    print(f"\nTheme distribution:")
    for k, v in sorted(stats.items()):
        print(f"  {k}: {v}")

    added_only = [c for c in changes if c["added"] and not c["removed"]]
    print(f"\nRows with supplemented themes: {len(added_only)}")

    # Summarize by type
    add_sol = sum(1 for c in added_only if "解决方案设计与评审" in c["added"])
    add_bid = sum(1 for c in added_only if "投标策略制定" in c["added"])
    add_tender = sum(1 for c in added_only if "招标策略制定" in c["added"])
    print(f"  +解决方案设计与评审: {add_sol} rows")
    print(f"  +招标策略制定: {add_tender} rows")
    print(f"  +投标策略制定: {add_bid} rows")

    print(f"\n=== Supplemented rows (for review) ===")
    for c in added_only:
        print(f"Row {c['row']}: [{c['old']}] -> [{c['new']}]")
        print(f"  {c['project'][:80]}")
        print(f"  {c['reason'][:250]}")
        print()

    # Log
    log_path = r"D:\数字分身\theme_changes.txt"
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"Theme Analysis Report\n====================\n\n")
        f.write(f"Per-theme coverage ({total} rows):\n")
        f.write(f"  解决方案设计与评审: {sol_count} ({100*sol_count//total}%)\n")
        f.write(f"  招标策略制定: {tender_count} ({100*tender_count//total}%)\n")
        f.write(f"  投标策略制定: {bid_count} ({100*bid_count//total}%)\n\n")
        f.write(f"Rows with supplemented themes: {len(added_only)}\n")
        f.write(f"  +解决方案设计与评审: {add_sol}\n")
        f.write(f"  +招标策略制定: {add_tender}\n")
        f.write(f"  +投标策略制定: {add_bid}\n\n")
        for c in added_only:
            f.write(f"Row {c['row']}: {c['project']}\n")
            f.write(f"  Old: {c['old']}\n")
            f.write(f"  New: {c['new']}\n")
            f.write(f"  Reason: {c['reason']}\n\n")
    print(f"Log: {log_path}")


if __name__ == "__main__":
    main()
